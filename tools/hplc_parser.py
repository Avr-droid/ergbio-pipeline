import os
import re
import csv
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# §15.1: Extended timepoint regex — T0, T24h, T96+h, T96, 24h, 96+h
TIMEPOINT_REGEX = re.compile(r'^T?\d+\+?h?$', re.IGNORECASE)

# §15.1: Row labels auto-flagged as controls (case-insensitive substring)
CONTROL_LABELS = {"abiotic", "blank", "control", "novo only", "no cell", "media only"}


def _is_control(label: str) -> bool:
    """Return True if this row should be auto-flagged as a control."""
    low = label.strip().lower()
    return any(kw in low for kw in CONTROL_LABELS)


def _find_timepoint_header_row(rows):
    """
    Scan rows to find the one that contains timepoint column headers.
    At least 2 cells must match TIMEPOINT_REGEX.
    """
    for i, row in enumerate(rows):
        matches = sum(1 for cell in row if cell and TIMEPOINT_REGEX.match(str(cell).strip()))
        if matches >= 2:
            return i
    return None


def _extract_analyte_values(rows, header_idx, analyte_name):
    """
    Find row(s) for a specific analyte and return {timepoint: value} dict.
    If multiple rows match (e.g. mean + stdev), takes the first.
    """
    header_row = rows[header_idx]
    for row in rows[header_idx + 1:]:
        if not row:
            continue
        label = str(row[0]).strip().lower() if row[0] else ""
        if analyte_name.lower() in label:
            values = {}
            for j, cell in enumerate(header_row):
                if cell and TIMEPOINT_REGEX.match(str(cell).strip()):
                    raw = row[j] if j < len(row) else None
                    try:
                        values[str(cell).strip()] = float(raw) if raw not in (None, "", "None") else None
                    except (ValueError, TypeError):
                        values[str(cell).strip()] = None
            return values
    return {}


def _extract_condition_rows(rows, header_idx, existing_conditions=None):
    """
    §15.1: Extract per-condition rows from the HPLC sheet.

    Each data row (below header) is treated as one condition.
    Rows are matched to `existing_conditions` by name (case-insensitive).
    Auto-flags control rows based on CONTROL_LABELS.

    Returns list of dicts:
        name, is_control, timepoints, t0_value, final_value, matched_condition
    """
    header_row = rows[header_idx]
    timepoint_cols = [
        (j, str(cell).strip())
        for j, cell in enumerate(header_row)
        if cell and TIMEPOINT_REGEX.match(str(cell).strip())
    ]

    # Build lookup for existing conditions (case-insensitive)
    cond_lookup = {}
    if existing_conditions:
        for c in existing_conditions:
            cond_lookup[c.get("name", "").strip().lower()] = c

    results = []
    for row in rows[header_idx + 1:]:
        if not row or not any(str(c).strip() for c in row):
            continue
        label = str(row[0]).strip() if row[0] else ""
        if not label:
            continue

        # Skip secondary stat rows (stdev, sem, cv%)
        low = label.lower()
        if any(kw in low for kw in ("stdev", "sd", "sem", "cv", "std dev", "%cv", "(g/l)", "(g/kg)")):
            continue

        ctrl = _is_control(label)
        matched = cond_lookup.get(label.lower())

        tp_values = {}
        for j, tp_label in timepoint_cols:
            raw = row[j] if j < len(row) else None
            try:
                tp_values[tp_label] = float(raw) if raw not in (None, "", "None") else None
            except (ValueError, TypeError):
                tp_values[tp_label] = None

        tp_labels = [tp for _, tp in timepoint_cols]
        t0     = tp_values.get(tp_labels[0])  if tp_labels else None
        final  = tp_values.get(tp_labels[-1]) if tp_labels else None

        results.append({
            "name":               label,
            "is_control":         ctrl,
            "timepoints":         tp_values,
            "t0_value":           t0,
            "final_value":        final,
            "matched_condition":  matched.get("name") if matched else None,
        })

    return results


def _rows_from_xlsx(file_path):
    """Load rows from an xlsx file, preferring a 'CBP' sheet."""
    if not HAS_OPENPYXL:
        return None, "openpyxl not installed — run: pip install openpyxl"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "cbp" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    return rows, None


def _rows_from_csv(file_path):
    try:
        rows = []
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
        return rows, None
    except FileNotFoundError:
        return None, f"File not found: {file_path}"
    except Exception as e:
        return None, f"Could not read file: {e}" 


def parse_hplc_file(file_path: str, existing_conditions=None) -> dict:
    """
    Parse an HPLC output file (.xlsx or .csv) for bioprocess data.

    §15.1 compliance:
    - Extended timepoint regex (T?\\d+\\+?h?)
    - Auto-flags control rows by name substring
    - Matches rows to existing_conditions (case-insensitive)

    Returns:
        success          bool
        timepoints       list of timepoint labels
        ethanol          dict {tp: g/L}
        glucose          dict {tp: g/L}
        xylose           dict {tp: g/L}
        ethanol_t0       value at first tp
        ethanol_final    value at last tp
        glucose_final    value at last tp
        xylose_final     value at last tp
        t0_values        dict {condition_name: t0 ethanol value}  (for condition table auto-fill)
        final_values     dict {condition_name: final ethanol value}
        condition_rows   list of per-condition dicts (name, is_control, t0_value, final_value, matched_condition)
        controls         list of control condition names auto-detected
        error            (only if success=False)
    """
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        rows, err = _rows_from_xlsx(file_path)
    elif ext == ".csv":
        rows, err = _rows_from_csv(file_path)
    else:
        return {"success": False, "error": f"Unsupported file type: {ext}"}

    if err:
        return {"success": False, "error": err}
    if not rows:
        return {"success": False, "error": "File appears to be empty"}

    header_idx = _find_timepoint_header_row(rows)
    if header_idx is None:
        return {"success": False,
                "error": "Could not find timepoint header row — expected ≥2 columns like T0, T24h, T96h"}

    header_row = rows[header_idx]
    timepoints = [str(cell).strip() for cell in header_row
                  if cell and TIMEPOINT_REGEX.match(str(cell).strip())]

    ethanol = _extract_analyte_values(rows, header_idx, "ethanol")
    glucose = _extract_analyte_values(rows, header_idx, "glucose")
    xylose  = _extract_analyte_values(rows, header_idx, "xylose")

    # Per-condition rows (used for CBP condition table auto-fill)
    cond_rows = _extract_condition_rows(rows, header_idx, existing_conditions)
    t0_values    = {c["name"]: c["t0_value"]    for c in cond_rows}
    final_values = {c["name"]: c["final_value"] for c in cond_rows}
    controls     = [c["name"] for c in cond_rows if c["is_control"]]

    first_tp = timepoints[0]  if timepoints else None
    last_tp  = timepoints[-1] if timepoints else None

    return {
        "success":         True,
        "timepoints":      timepoints,
        "ethanol":         ethanol,
        "glucose":         glucose,
        "xylose":          xylose,
        "ethanol_t0":      ethanol.get(first_tp) if first_tp else None,
        "ethanol_final":   ethanol.get(last_tp)  if last_tp  else None,
        "glucose_final":   glucose.get(last_tp)  if last_tp  else None,
        "xylose_final":    xylose.get(last_tp)   if last_tp  else None,
        "t0_values":       t0_values,
        "final_values":    final_values,
        "condition_rows":  cond_rows,
        "controls":        controls,
    }
